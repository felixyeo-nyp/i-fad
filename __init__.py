# Flask-related imports #
import flask_mail
from flask import Flask, render_template, request, redirect, url_for, flash, Response, jsonify, send_file, session, \
    current_app
import socket
import struct
from scapy.all import *
from scapy.layers.inet import IP, TCP
from sympy import false
from wtforms import Form, StringField, RadioField, SelectField, TextAreaField, validators, ValidationError, PasswordField
import sys
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user,current_user
import shelve, re
from flask_wtf import FlaskForm, CSRFProtect
from wtforms.validators import email
from Forms import configurationForm, emailForm, LoginForm, RegisterForm,updatepasswordForm, MFAForm, FeedbackForm, updateemailrole, forgetpassword , ipForm, DeleteForm
from flask_mail import Mail, Message
import random
import secrets
import webbrowser
# Object detection & processing-related imports #
import torch
import torchvision
import cv2
import torchvision.models as models
from torchvision.models.detection.faster_rcnn import FastRCNNPredictor

# Datetime-related imports #
import datetime, time
import threading
from datetime import datetime, timedelta
from flask_apscheduler import APScheduler
import numpy as np
import pytz

# File-reader-related imports #
import openpyxl
from openpyxl.styles import Font

# OOP-related imports #
from OOP import *
import os

# Security-related imports #
from functools import wraps
from flask import abort

# System-related imports #
import socket
import psutil
import uuid

# Back-end codes for object detection & processing #
if torch.cuda.is_available():
    print('you are using gpu to process the video camera')
else:
    print('no gpu is found in this python environment. using cpu to process')

class FreshestFrame(threading.Thread):
    def __init__(self, capture, name='FreshestFrame'):
        if not capture or not capture.isOpened():
            raise ValueError("Capture device is not opened.")

        self.capture = capture
        self.condition = threading.Condition()
        self.stop_event = camera_stop_event
        self.is_running = False
        self.frame = None
        self.pellets_num = 0
        self.callback = None
        super().__init__(name=name, daemon=True)
        self.start()

    def stop(self, timeout=None):
        self.stop_event.set()
        self.is_running = False
        self.join(timeout=timeout)
        if self.capture and self.capture.isOpened():
            self.capture.release()

    def run(self):
        self.is_running = True
        print("[FreshestFrame] Thread started.")
        try:
            while not self.stop_event.is_set():
                try:
                    if not self.capture.isOpened():
                        print("[WARNING] Capture device closed.")
                        break

                    rv, img = self.capture.read()
                    if not rv or img is None:
                        print("[WARNING] Frame read failed, retrying...")
                        time.sleep(1)
                        continue

                    with self.condition:
                        self.frame = img
                        self.pellets_num += 1
                        self.condition.notify_all()

                    if self.callback:
                        try:
                            self.callback(img)
                        except Exception as cb_err:
                            print(f"[ERROR] Callback failed: {cb_err}")
                except cv2.error as cv_err:
                    print(f"[OpenCV ERROR] {cv_err}")
                    traceback.print_exc()
                    time.sleep(1)
                except Exception as e:
                    print(f"[ERROR] Unexpected error in FreshestFrame thread: {e}")
                    traceback.print_exc()
                    time.sleep(1)
        finally:
            print("[INFO] Releasing capture resource.")
            if self.capture and self.capture.isOpened():
                self.capture.release()
            print("[FreshestFrame] Thread stopped cleanly.")

    def read(self, wait=True, sequence_number=None, timeout=None):
        with self.condition:
            if wait:
                if sequence_number is None:
                    sequence_number = self.pellets_num + 1

                if sequence_number < 1:
                    sequence_number = 1

                rv = self.condition.wait_for(
                    lambda: self.pellets_num >= sequence_number,
                    timeout=timeout
                )
                if not rv:
                    print("[WARNING] Frame wait timeout, returning latest available frame.")
                    return (self.pellets_num, self.frame)

            return (self.pellets_num, self.frame)

# define the id "1" for pellets
# do note that in the pth file, the pellet id also is 1
class_labels = {
    1: 'Pellets',
}

# pth file where you have defined on roboflow
model_path = './best_model5.pth'
latest_processed_frame = None  # Stores the latest processed frame
stop_event = threading.Event()  # Event to stop threads gracefully
camera_stop_event = threading.Event()
freshest_frame = None

# Initialize variables for feeding logic
feeding = False
feeding_timer = None
showing_timer = None
line_chart_timer = None
object_count = {1: 0}
frame_data = {
    'object_count': {1: 0},  # Initialize with default values for object count
    'bounding_boxes': []  # List to store bounding boxes for the current frame
}

# Initialize locks for shared variables
latest_processed_frame_lock = threading.Lock()
frame_data_lock = threading.Lock()
object_count_lock = threading.Lock()
freshest_frame_lock = threading.Lock()
shelve_lock = threading.Lock()

def create_model(num_classes, pretrained=False, coco_model=False):
    if pretrained:
        weights = torchvision.models.detection.FasterRCNN_ResNet50_FPN_Weights.DEFAULT
    else:
        weights = None

    model = torchvision.models.detection.fasterrcnn_resnet50_fpn(weights=weights)

    if not coco_model:
        in_features = model.roi_heads.box_predictor.cls_score.in_features
        model.roi_heads.box_predictor = FastRCNNPredictor(in_features, num_classes)

    return model

# Function to load the custom-trained model from the .pth file
def load_model(model_path, num_classes):
    model = create_model(num_classes=num_classes, pretrained=False, coco_model=False)
    model.roi_heads.detections_per_img = 500
    checkpoint = torch.load(model_path, map_location=device)
    model.load_state_dict(checkpoint['model_state_dict'])

    return model

# Assume these methods to load model and settings are defined
device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
print(device)
model = load_model(model_path, num_classes=2)
model.to(device)
model.eval()

def capture_frames():
    while not camera_stop_event.is_set():
        try:
            with shelve.open('IP.db', 'c') as ip_db:
                cam = ip_db.get('IP', {})
                ip = cam.get('camera_ip')
                user = cam.get('amcrest_username')
                pw = cam.get('amcrest_password')

                if not all([ip, user, pw]):
                    print("Camera config incomplete, waiting...")
                    time.sleep(5)
                    continue  # Try again in a few seconds

                rtsp_url = f"rtsp://{user}:{pw}@{ip}:554/cam/realmonitor?channel=1&subtype=0"
                print(f"Connecting to RTSP stream: {rtsp_url}")
        except Exception as e:
            print(f"Error reading IP.db: {e}")
            time.sleep(5)
            continue

        cap = cv2.VideoCapture(rtsp_url)

        if not cap.isOpened():
            print("Warning: Cannot open video stream. Retrying in 10s...")
            cap.release()
            time.sleep(10)
            continue

        try:
            global freshest_frame
            with freshest_frame_lock:
                freshest_frame = FreshestFrame(cap)
        except Exception as e:
            print(f"[ERROR] Could not start FreshestFrame thread: {e}")
            cap.release()
            continue

        global latest_processed_frame
        while not camera_stop_event.is_set():
            if not cap.isOpened():
                break

            sequence_num, frame = freshest_frame.read(wait=True)
            if frame is not None:
                with latest_processed_frame_lock:
                    latest_processed_frame = frame

            time.sleep(0.03)

        cap.release()

def validate_config_thread():
    global latest_valid_settings, latest_set_ip_settings
    print("Starting validation thread for settings and IP...")

    fallback_confidence_rate = 60
    fallback_ip = {
        "source": "192.168.0.65",
        "destination": "192.168.0.18",
        "camera_ip": "192.168.0.52",
        "amcrest_username": "admin",
        "amcrest_password": "fyp2025Fish34535"
    }
    fallback_email = Email('iatfadteam@gmail.com', 'iatfadteam@gmail.com', 'pmtu cilz uewx xqqi', 3)

    while True:
        try:
            # Check IP.db
            ip_config_updated = False
            with shelve_lock:
                with shelve.open('IP.db', 'c') as ip_db:
                    ip_data = ip_db.get("IP", {})
                    required_keys = ["source", "destination", "camera_ip", "amcrest_username", "amcrest_password"]

                    if not all(k in ip_data and ip_data[k] for k in required_keys):
                        ip_db["IP"] = latest_set_ip_settings if latest_set_ip_settings else fallback_ip
                        ip_config_updated = True
                        print("[VALIDATE] Fallback IP configuration written.")

            # Check settings.db
            with shelve_lock:
                with shelve.open('settings.db', 'c') as settings_db:
                    if 'Confidence_Rate' not in settings_db:
                        settings_db['Confidence_Rate'] = fallback_confidence_rate
                        print("[VALIDATE] Fallback Confidence_Rate initialized.")

                    if 'Port' not in settings_db:
                        settings_db['Port'] = 53101
                        print("[VALIDATE] Missing Port generated.")

                    if 'Generate_Status' not in settings_db:
                        settings_db['Generate_Status'] = False
                        print("[VALIDATE] Generate_Status set to False.")

                    if 'Email_Data' not in settings_db:
                        settings_db['Email_Data'] = {'Email_Info': fallback_email}
                        print("[VALIDATE] Fallback Email_Data written.")

                    if ip_config_updated or 'syn_ack_seq' not in settings_db or 'syn_ack_ack' not in settings_db:
                        try:
                            ip_data = None
                            with shelve_lock:
                                with shelve.open('IP.db', 'r') as ip_db:
                                    ip_data = ip_db.get("IP", {})

                            source = ip_data.get("source")
                            destination = ip_data.get("destination")

                            port = settings_db.get('Port')
                            if source and destination and port:
                                start_syn_packet_thread(source, destination, port, 50000)
                                send_udp_packet(source, "255.255.255.255", 60000, b"\x00\x00\x00\x00\x00")
                                print("[VALIDATE] SYN/ACK packets sent after IP config.")
                            else:
                                print("[VALIDATE WARNING] Could not send SYN/ACK packets due to missing source/destination/port.")

                        except Exception as e:
                            print(f"[VALIDATE ERROR] Failed to generate SYN/ACK packets: {e}")

        except Exception as e:
            print(f"[VALIDATE ERROR] {e}")

        time.sleep(2)

def video_processing_loop():
    global freshest_frame, object_count, frame_data, latest_processed_frame
    while True:
        try:
            with shelve_lock:
                with shelve.open('settings.db', 'r') as db:
                    confidence = float(db.get('Confidence_Rate', 60)) / 100
        except Exception as e:
            print(f"[ERROR] Failed to read confidence rate from settings.db: {e}")
            time.sleep(1)
            continue

        with freshest_frame_lock:
            if freshest_frame is None:
                time.sleep(0.1)
                continue
            cnt, frame = freshest_frame.read(sequence_number=object_count[1] + 1)
            if frame is None:
                time.sleep(0.1)
                continue

        # Inference
        img_tensor = torchvision.transforms.ToTensor()(frame).unsqueeze(0).to(device)
        with torch.no_grad():
            predictions = model(img_tensor)

        temp_object_count = {1: 0}
        bounding_boxes = []
        for i in range(len(predictions[0]['labels'])):
            label = predictions[0]['labels'][i].item()
            if label == 1 and predictions[0]['scores'][i].item() > confidence:
                box = predictions[0]['boxes'][i].cpu().numpy().astype(int)
                temp_object_count[label] += 1
                bounding_boxes.append((box, label, predictions[0]['scores'][i].item()))
        
        with object_count_lock:
            object_count[1] = temp_object_count[1]

        try:
            with shelve_lock:
                with shelve.open('currentcount.db', 'c') as db2:
                    db2['object_count'] = temp_object_count[1]
        except Exception as e:
            print(f"[ERROR] Failed to write to currentcount.db: {e}")

        with frame_data_lock:
            frame_data['object_count'] = temp_object_count
            frame_data['bounding_boxes'] = bounding_boxes
        with latest_processed_frame_lock:
            latest_processed_frame = frame

def feeding_scheduler_loop():
    global feeding, feed_start_time, total_count
    feeding = False
    feed_start_time = None
    total_count = 0
    current_feeding_session = None
    feed_cycle_start_time = None
    is_feeding_phase = False
    feeding_amount_storage = 0
    current_cycle = 0
    port = server_isn = server_ack = source_ip = destination_ip = None

    while True:
        user_data = get_active_feeding_user()
        user_uuid = user_data.get("uuid")
        config = user_data.get("config")

        if not user_uuid or not config:
            time.sleep(1)
            continue

        tz = pytz.timezone("Asia/Singapore")
        now = datetime.now(tz)

        try:
            # Get feeding times
            morning_feed_1 = config.get("user_morning_feed_1", "")
            morning_feed_2 = config.get("user_morning_feed_2", "")
            evening_feed_1 = config.get("user_evening_feed_1", "")
            evening_feed_2 = config.get("user_evening_feed_2", "")
            
            # Get feeding parameters
            feeding_duration = int(config.get("user_minutes", 0)) * 60
            check_interval = int(config.get("user_interval_seconds", 2))
            user_pellets = int(config.get("user_pellets", 0))
            feeding_threshold = int(config.get("user_feeding_threshold", 0))
            pellets_per_second = int(config.get("user_pellets_per_second", 1))
            
            required_feed_time = user_pellets / pellets_per_second
            required_feed_time = math.ceil(required_feed_time)
            available_check_time = feeding_duration - required_feed_time
            
            if available_check_time <= 0:
                max_checks = 0
                feed_per_cycle = required_feed_time
                check_per_cycle = 0
                total_cycles = 1
                print(f"[CALC] No check time available. Continuous feeding for {feed_per_cycle}s")
            else:
                max_checks = available_check_time // check_interval
                
                if max_checks == 0:
                    max_checks = 1
                    feed_per_cycle = required_feed_time
                    check_per_cycle = available_check_time
                    total_cycles = 1
                else:
                    feed_per_cycle = math.ceil(required_feed_time / max_checks)
                    check_per_cycle = check_interval
                    total_cycles = max_checks

            pellets_per_feed_cycle = feed_per_cycle * pellets_per_second
            
            scheduled_times = []
            if morning_feed_1 and morning_feed_1 != "N/A":
                scheduled_times.append((now.replace(hour=int(morning_feed_1[:2]), minute=int(morning_feed_1[2:]), second=0, microsecond=0), "morning_feed_1"))
            if morning_feed_2 and morning_feed_2 != "N/A":
                scheduled_times.append((now.replace(hour=int(morning_feed_2[:2]), minute=int(morning_feed_2[2:]), second=0, microsecond=0), "morning_feed_2"))
            if evening_feed_1 and evening_feed_1 != "N/A":
                scheduled_times.append((now.replace(hour=int(evening_feed_1[:2]), minute=int(evening_feed_1[2:]), second=0, microsecond=0), "evening_feed_1"))
            if evening_feed_2 and evening_feed_2 != "N/A":
                scheduled_times.append((now.replace(hour=int(evening_feed_2[:2]), minute=int(evening_feed_2[2:]), second=0, microsecond=0), "evening_feed_2"))

        except Exception as e:
            print(f"[ERROR] Invalid feeding config: {e}")
            time.sleep(1)
            continue

        with object_count_lock:
            current_count = object_count[1]

        # Check if feeding should start
        if not feeding:
            for scheduled_time, session_name in scheduled_times:
                if scheduled_time.strftime('%H:%M:%S') <= now.strftime('%H:%M:%S') <= str((scheduled_time + timedelta(seconds=5)).strftime('%H:%M:%S')):
                    print(f"[FEED TRIGGER] Starting feeding session {session_name} at {now.strftime('%H:%M:%S')}")
                    current_feeding_session = session_name
                    # Start camera threads for pellets detection
                    try:
                        start_camera_threads()
                    except Exception as e:
                        print(f"[ERROR] Failed to start camera threads: {e}")
                        continue
                    
                    try:
                        with shelve_lock:
                            with shelve.open('IP.db', 'r') as ip_db, shelve.open('settings.db', 'r') as db:
                                port = db.get('Port')
                                server_isn = db.get('syn_ack_seq')
                                server_ack = db.get('syn_ack_ack')
                                ip_data = ip_db.get('IP', {})
                                source_ip = ip_data.get('source')
                                destination_ip = ip_data.get('destination')

                        feeding = True
                        feed_start_time = time.time()
                        feed_cycle_start_time = time.time()
                        is_feeding_phase = True
                        total_count = 0
                        feeding_amount_storage = 0
                        current_cycle = 1
                        
                        # Start first feeding phase
                        success = start_send_manual_feed(port, server_isn, server_ack, source_ip, destination_ip)
                        if success:
                            print(f"[FEED START] Started feeding phase of cycle {current_cycle}/{total_cycles} ({feed_per_cycle}s)")
                            # Add pellets for this feeding cycle to storage
                            feeding_amount_storage += pellets_per_feed_cycle
                            print(f"[COUNT] Added {pellets_per_feed_cycle}g to storage. Total: {feeding_amount_storage}g")
                        else:
                            print(f"[FEED ERROR] Failed to start feeding")
                            
                    except Exception as e:
                        print(f"[ERROR] Failed to retrieve database settings: {e}")
                    break

        # If in feeding mode, manage feed/check cycles
        if feeding and feed_start_time is not None and feed_cycle_start_time is not None:
            cycle_elapsed = time.time() - feed_cycle_start_time
            
            with object_count_lock:
                total_count += object_count[1]

            # Feeding phase
            if is_feeding_phase and cycle_elapsed >= feed_per_cycle:
                # End feeding phase, start check phase (if check time > 0)
                try:
                    with shelve_lock:
                        with shelve.open('IP.db', 'r') as ip_db, shelve.open('settings.db', 'r') as db:
                            port = db.get('Port')
                            server_isn = db.get('syn_ack_seq')
                            server_ack = db.get('syn_ack_ack')
                            ip_data = ip_db.get('IP', {})
                            source_ip = ip_data.get('source')
                            destination_ip = ip_data.get('destination')
                    
                    success = stop_send_manual_feed(port, server_isn, server_ack, source_ip, destination_ip)
                    if success:
                        print(f"[FEED PAUSE] Completed feeding cycle {current_cycle}/{total_cycles}")
                    else:
                        print(f"[FEED ERROR] Failed to pause feeding")
                    
                    # If there's no check time or this is the last cycle, end session
                    if check_per_cycle == 0 or current_cycle >= total_cycles:
                        print(f"[FEED END] All feeding cycles completed.")
                        # Stop camera threads
                        try:
                            stop_camera_threads()
                        except Exception as e:
                            print(f"[ERROR] Failed to stop camera threads: {e}")
                            continue
                        feeding = False
                        feed_start_time = None
                        feed_cycle_start_time = None
                        if current_feeding_session:
                            session_times = {
                                'morning_feed_1': morning_feed_1,
                                'morning_feed_2': morning_feed_2,
                                'evening_feed_1': evening_feed_1,
                                'evening_feed_2': evening_feed_2
                            }
                            feeding_time_str = session_times.get(current_feeding_session, "")
                            save_chart_data(feeding_amount_storage, current_feeding_session, feeding_time_str)
                            print(f"[FEED END] Session {current_feeding_session} ended with {feeding_amount_storage}g dispensed.")
                            feeding_amount_storage = 0
                            current_feeding_session = None
                    else:
                        is_feeding_phase = False
                        
                except Exception as e:
                    print(f"[ERROR] Failed to pause feed: {e}")

            # Check phase
            elif not is_feeding_phase and cycle_elapsed >= (feed_per_cycle + check_per_cycle):
                # Check phase completed, decide whether to continue
                with object_count_lock:
                    current_count = object_count[1]

                feeding_threshold_with_buffer = feeding_threshold + 5  # Add a buffer for unexpected counts (E.g. Water Reflection, Camera Glitches, etc.)

                print(f"[FEED CHECK] Cycles: {current_cycle}/{total_cycles}, Count: {current_count}/{feeding_threshold_with_buffer}")
                if current_cycle >= total_cycles:
                    # End feeding session
                    print(f"[FEED END] Feeding session completed. Cycles: {current_cycle}/{total_cycles}, Count: {current_count}/{feeding_threshold_with_buffer}")
                    # Stop camera threads
                    try:
                        stop_camera_threads()
                    except Exception as e:
                        print(f"[ERROR] Failed to stop camera threads: {e}")
                        continue
                    feeding = False
                    feed_start_time = None
                    feed_cycle_start_time = None
                    if current_feeding_session:
                        session_times = {
                            'morning_feed_1': morning_feed_1,
                            'morning_feed_2': morning_feed_2,
                            'evening_feed_1': evening_feed_1,
                            'evening_feed_2': evening_feed_2
                        }
                        feeding_time_str = session_times.get(current_feeding_session, "")
                        save_chart_data(feeding_amount_storage, current_feeding_session, feeding_time_str)
                        print(f"[FEED END] Session {current_feeding_session} ended with {feeding_amount_storage}g dispensed.")
                        feeding_amount_storage = 0
                        current_feeding_session = None
                else:
                    # Continue to next cycle if pellets are below threshold
                    if current_count < feeding_threshold_with_buffer:
                        try:
                            with shelve_lock:
                                with shelve.open('IP.db', 'r') as ip_db, shelve.open('settings.db', 'r') as db:
                                    port = db.get('Port')
                                    server_isn = db.get('syn_ack_seq')
                                    server_ack = db.get('syn_ack_ack')
                                    ip_data = ip_db.get('IP', {})
                                    source_ip = ip_data.get('source')
                                    destination_ip = ip_data.get('destination')
                            
                            current_cycle += 1
                            success = start_send_manual_feed(port, server_isn, server_ack, source_ip, destination_ip)
                            if success:
                                print(f"[FEED CONTINUE] Started feeding cycle {current_cycle}/{total_cycles} ({feed_per_cycle}s)")
                                # Add pellets for this feeding cycle to storage
                                feeding_amount_storage += pellets_per_feed_cycle
                                print(f"[COUNT] Added {pellets_per_feed_cycle}g to storage. Total: {feeding_amount_storage}g")
                            else:
                                print(f"[FEED ERROR] Failed to continue feeding")
                                
                            feed_cycle_start_time = time.time()
                            is_feeding_phase = True
                            
                        except Exception as e:
                            print(f"[ERROR] Failed to continue feed: {e}")
                    else:
                        # Pellet level sufficient, skip this feeding cycle and wait for next check
                        print(f"[FEED SKIP] Pellet threshold reached. Skipping cycle {current_cycle + 1}, waiting for next check.")
                        current_cycle += 1
                        feed_cycle_start_time = time.time()
                        is_feeding_phase = False  # Go directly to next check phase

            # Safety check - if total duration exceeds feeding_duration, force stop
            if feed_start_time is not None:
                total_elapsed = time.time() - feed_start_time
                if total_elapsed >= feeding_duration:
                    print(f"[FEED TIMEOUT] Maximum duration exceeded. Force stopping.")
                    try:
                        with shelve_lock:
                            with shelve.open('IP.db', 'r') as ip_db, shelve.open('settings.db', 'r') as db:
                                port = db.get('Port')
                                server_isn = db.get('syn_ack_seq')
                                server_ack = db.get('syn_ack_ack')
                                ip_data = ip_db.get('IP', {})
                                source_ip = ip_data.get('source')
                                destination_ip = ip_data.get('destination')
                        
                        success = stop_send_manual_feed(port, server_isn, server_ack, source_ip, destination_ip)
                        if success:
                            print("[FEED TIMEOUT] Feed stopped successfully")
                        else:
                            print("[FEED TIMEOUT ERROR] Failed to stop feed")
                            
                    except Exception as e:
                        print(f"[ERROR] Failed to force stop feed: {e}")

                    # Stop camerat threads
                    try:
                        stop_camera_threads()
                    except Exception as e:
                        print(f"[ERROR] Failed to stop camera threads: {e}")
                        continue
                        
                    feeding = False
                    feed_start_time = None
                    feed_cycle_start_time = None
                    if current_feeding_session:
                        session_times = {
                            'morning_feed_1': morning_feed_1,
                            'morning_feed_2': morning_feed_2,
                            'evening_feed_1': evening_feed_1,
                            'evening_feed_2': evening_feed_2
                        }
                        feeding_time_str = session_times.get(current_feeding_session, "")
                        save_chart_data(feeding_amount_storage, current_feeding_session, feeding_time_str)
                        print(f"[FEED TIMEOUT] Session {current_feeding_session} ended with timeout at {feeding_amount_storage}g dispensed.")
                        feeding_amount_storage = 0
                        current_feeding_session = None

        time.sleep(1)

def save_chart_data(feeding_amount_storage, feeding_session, feeding_time_str):
    """
    Save chart data based on the feeding_amount_storage (actual pellets dispensed during feeding cycles)
    """
    try:
        # Convert session name to display format
        session_display = {
            'morning_feed_1': 'Morning Feed 1',
            'morning_feed_2': 'Morning Feed 2', 
            'evening_feed_1': 'Evening Feed 1',
            'evening_feed_2': 'Evening Feed 2'
        }
        
        session = session_display.get(feeding_session, feeding_session)

        with shelve_lock:
            with shelve.open('mock_chart_data.db', 'c') as db:
                date_str = datetime.today().strftime("%Y-%m-%d")

                if date_str not in db:
                    db[date_str] = {}

                day_data = db[date_str]
                day_data[session] = day_data.get(session, 0) + int(feeding_amount_storage)
                day_data['Total'] = day_data.get('Total', 0) + int(feeding_amount_storage)

                db[date_str] = day_data

        print(f"[SAVE] {session} session ({feeding_time_str}) recorded with {feeding_amount_storage} pellets.")
        
    except Exception as e:
        print(f"[ERROR] Failed to save feeding data: {e}")

def generate_frames():
    global latest_processed_frame, frame_data
    count = 1
    while not camera_stop_event.is_set():
        if count // 60 == 0:
            db = shelve.open('settings.db', 'c')
            if not db.get('Generate_Status', True):
                print("stopped generating")
                break
            db.close()

        with latest_processed_frame_lock:
            if latest_processed_frame is None:
                # Create a black frame fallback if no frame available
                black_frame = np.zeros((480, 640, 3), dtype=np.uint8)
                frame_to_use = black_frame
            else:
                frame_to_use = latest_processed_frame.copy()

        # Overlay info only if it's not the black fallback frame
        if latest_processed_frame is not None:
            with frame_data_lock:
                for label, count in frame_data['object_count'].items():
                    text = f'{class_labels[label]} Count: {count}'
                    text_size = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, 1.2, 2)[0]
                    text_position = (frame_to_use.shape[1] - text_size[0] - 10, 30 * (label + 1))
                    cv2.putText(frame_to_use, text, text_position, cv2.FONT_HERSHEY_SIMPLEX, 1.2, (200, 255, 255), 2)

                for box, label, score in frame_data['bounding_boxes']:
                    cv2.rectangle(frame_to_use, (box[0], box[1]), (box[2], box[3]), (0, 255, 0), 2)
                    cv2.putText(frame_to_use, f'{class_labels[label]}: {score:.2f}', (box[0], box[1] - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 1.4, (0, 255, 0), 2)
        try:
            ret, jpeg = cv2.imencode('.jpg', frame_to_use)
        except Exception as e:
            print("[Generate Frames] Encoding error:", e)
        if ret:
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + jpeg.tobytes() + b'\r\n\r\n')
        else:
            time.sleep(0.1)

        count += 1
        time.sleep(0.03)  # Adjust frame rate if necessary

# Web application #
app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key'
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
scheduler = APScheduler()
scheduler.init_app(app)
scheduler.start()
csrf = CSRFProtect(app)

# Initialize Flask-Mail
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USERNAME'] = 'iatfadteam@gmail.com'
app.config['MAIL_DEFAULT_SENDER'] = ('Admin', 'iatfadteam@gmail.com')
app.config['MAIL_PASSWORD'] = 'pmtu cilz uewx xqqi'
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
mail = Mail(app)

# Dictionaries #
#j = datetime.datetime.now()
# print(j)
Line_Chart_Data_dict = {}
Email_dict = {}

# User model
class User(UserMixin):
    def __init__(self, username, email, password, role, status="Active"):
        self.id = username  # Use username as ID for simplicity
        self.username = username
        self.email = email
        self.password = password
        self.role = role
        self.status = status

    def get_id(self):
        return str(self.username)

def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'role' not in session or session['role'] != role:
                username = session.get('username')
                if username:  # Ensure the user is logged in
                    with shelve.open('users.db', writeback=True) as db:
                        user = db.get(username)
                        if user and user["status"] == "Active":
                            user["status"] = "Breached"  # Change status to Breached
                            db[username] = user  # Save the updated user back to the database
                return redirect(url_for("breached"))  # Redirect to Breached page
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.route("/breached", methods=['GET'])
def breached():
    session.clear()
    return render_template("breached.html")

@app.errorhandler(404)
def page_not_found(error):
    return render_template('404.html'), 404

# User loader callback for Flask-Login
@login_manager.user_loader
def load_user(user_id):
    try:
        user_id = user_id.decode() if isinstance(user_id, bytes) else user_id

        with shelve.open('users.db') as db:
            user_data = db.get(user_id)
            if user_data:
                return User(
                    user_data['username'],
                    user_data['email'],
                    user_data['password'],
                    user_data['role'],
                    user_data.get('status', 'Active')
                )
    except Exception as e:
        app.logger.error(f"Error loading user {user_id}: {e}")
    return None

# Default route to redirect to login page
@app.route('/')
@login_required
def index():
    return redirect(url_for('logout'))

def seed_admin_account():
    with shelve.open('users.db', 'c') as db:
        has_admin = False
        for key, user_data in db.items():
            if user_data.get('role') == 'Admin':
                has_admin = True
                break

        if not has_admin:
            hashed = generate_password_hash('Password1!', method='pbkdf2:sha256')
            admin = User(
                username='admin',
                email='testproject064@gmail.com',
                password=hashed,
                role='Admin'
            )
            db['admin'] = {
                'uuid': str(uuid.uuid4()),
                'username': admin.username,
                'email': admin.email.lower(),
                'password': admin.password,
                'role': admin.role,
                'status': admin.status,
                'user_morning_feed_1': "",
                'user_morning_feed_2': "",
                'user_evening_feed_1': "",
                'user_evening_feed_2': "",
                'user_minutes': 0,
                'user_interval_seconds': 0,
                'user_pellets': 0,
                'user_feeding_threshold': 0,
                'user_pellet_size': 1,
                'user_pellets_per_second': 1
            }
            print("Admin account seeded successfully.")

seed_admin_account()
@app.route('/register', methods=['GET', 'POST'])
def register():
    form = RegisterForm()  # Create an instance of RegisterForm

    if form.validate_on_submit():
        username = form.username.data.strip()
        email = form.email.data.strip().lower()
        password = form.password.data
        role = form.role.data

        # Check if the username or email already exists in the database
        with shelve.open('users.db', 'c') as db:
            username_exists = username in db
            email_exists = any(user_data['email'].lower() == email for user_data in db.values())

            if username_exists or email_exists:
                flash('Username or email already in use', 'danger')
            else:
                user_uuid = str(uuid.uuid4())
                hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
                new_user = User(username, email, hashed_password, role, status="Active")
                db[username] = {
                    'uuid': user_uuid,
                    'username': new_user.username,
                    'email': new_user.email,
                    'password': new_user.password,
                    'role': new_user.role,
                    'status': new_user.status,
                    'user_morning_feed_1': "",
                    'user_morning_feed_2': "",
                    'user_evening_feed_1': "",
                    'user_evening_feed_2': "",
                    'user_minutes': 0,
                    'user_interval_seconds': 0,
                    'user_pellets': 0,
                    'user_feeding_threshold': 0,
                    'user_pellet_size': 1,
                    'user_pellets_per_second': 1
                }
                flash('You are now registered and can log in', 'success')
                return redirect(url_for('login'))

    return render_template('register.html', form=form, hide_sidebar=True)

@app.route('/register2', methods=['GET', 'POST'])
@login_required
@role_required('Admin')
def register2():
    form = RegisterForm()  # Create an instance of RegisterForm

    if form.validate_on_submit():
        username = form.username.data.strip()
        email = form.email.data.strip().lower()
        password = form.password.data
        role = form.role.data

        # Check if the username or email already exists in the database
        with shelve.open('users.db', 'c') as db:
            username_exists = username in db
            email_exists = any(user_data['email'].lower() == email for user_data in db.values())

            if username_exists or email_exists:
                    flash('Username or email already in use', 'danger')
            else:
                user_uuid = str(uuid.uuid4())
                hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
                new_user = User(username, email, hashed_password, role, status="Active")
                db[username] = {
                    'uuid': user_uuid,
                    'username': new_user.username,
                    'email': new_user.email,
                    'password': new_user.password,
                    'role': new_user.role,
                    'status': new_user.status,
                    'user_morning_feed_1': "",
                    'user_morning_feed_2': "",
                    'user_evening_feed_1': "",
                    'user_evening_feed_2': "",
                    'user_minutes': 0,
                    'user_interval_seconds': 0,
                    'user_pellets': 0,
                    'user_feeding_threshold': 0,
                    'user_pellet_size': 1,
                    'user_pellets_per_second': 1
                }
                return redirect(url_for('retrieve_users'))

    return render_template('register2.html', form=form)

def find_user(identifier: str):
    identifier = identifier.strip().lower()
    with shelve.open('users.db', 'r') as db:
        # 1) Username matching
        for uname in db:
            if uname.lower() == identifier:
                user = db[uname]
                user['username'] = uname
                return user

        # 2) Email matching
        for uname, udata in db.items():
            if udata.get('email', '').lower() == identifier:
                udata['username'] = uname
                return udata

    return None

@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        # ← now reads from form.identifier, which can be either
        identifier = form.identifier.data.strip().lower()
        password   = form.password.data

        user = find_user(identifier)
        if user and check_password_hash(user['password'], password):
            if user['status'] in ["Suspended", "Breached"]:
                flash(f"Your account is {user['status']}. Access denied.", "danger")
                return redirect(url_for('login'))

            session['username'] = user['username']
            session['email']    = user['email']
            session['role']     = user['role']
            with shelve.open('settings.db', writeback=True) as settings_db:
                settings_db['CurrentUserEmail'] = user['email']
            return redirect(url_for('mfa_verify'))

        flash('Invalid login credentials', 'danger')

    return render_template('login.html', form=form, hide_sidebar=True)


# send OTP (Login)
from datetime import datetime, timezone

def send_mfa_code():
    # 1) generate code & timestamps
    code = str(secrets.randbelow(900_000) + 100_000)
    print("The MFA Code: ",code)
    now = datetime.now(timezone.utc)
    session['mfa_code']       = code
    session['mfa_sent_at']    = now.isoformat()
    session['last_resend_at'] = now.isoformat()
    current_app.logger.debug(f"[DEBUG] Login MFA code is: {code}")

    # 2) build plain-text body
    text_body = (
        "I@FAD MFA Verification\n\n"
        f"Your 6-digit code is: {code}\n"
        "It expires in 60 seconds.\n\n"
        "If you didn’t request this, ignore this email."
    )

    # 3) load & format static HTML
    html_path = current_app.root_path + '/templates/email/login_mfa.html'
    with open(html_path, 'r') as f:
        html_body = f.read().format(
            code=code,
            expires=60,
            year=now.year
        )

    # 4) construct Message and send
    try:
        msg = flask_mail.Message(
            subject="I@FAD Code",
            sender=current_app.config['MAIL_DEFAULT_SENDER'],
            recipients=[ session.get('email') ],
            body=text_body,
            html=html_body
        )
        mail.send(msg)
        return True

    except Exception as e:
        current_app.logger.error(f"[ERROR] Failed to send MFA email: {e}")
        return False



@app.route('/mfa-verify', methods=['GET', 'POST'])
def mfa_verify():
    # 1) Ensure user just logged in
    if 'username' not in session:
        return redirect(url_for('login'))

    form     = MFAForm()
    now      = datetime.now(timezone.utc)
    sent_iso = session.get('mfa_sent_at')

    # 2) Check if existing code is expired
    expired = True
    if sent_iso:
        sent_at = datetime.fromisoformat(sent_iso)
        expired = (now - sent_at).total_seconds() > 60

    # 3) On GET: generate & send when missing or expired
    if request.method == 'GET':
        if not sent_iso or expired:
            if not send_mfa_code():
                flash(
                    "Unable to send authentication code right now. "
                    "Please try logging in again later.",
                    'danger'
                )
                return redirect(url_for('login'))

            # only flash the “code sent” info on the very first GET
            if not sent_iso:
                flash('An authentication code has been sent to your email.', 'info')
            # if expired, we’re re-sending silently (the POST already showed “Code expired…”)

    # 4) On POST: validate the code
    if form.validate_on_submit():
        entered = form.code.data
        sent_iso = session.get('mfa_sent_at')

        if not sent_iso:
            flash('No code found. Please wait for the new one.', 'danger')
            return redirect(url_for('mfa_verify'))

        sent_at = datetime.fromisoformat(sent_iso)
        now     = datetime.now(timezone.utc)

        if (now - sent_at).total_seconds() > 60:
            flash('Code expired. A new code has been sent.', 'danger')
            return redirect(url_for('mfa_verify'))

        if entered == session.get('mfa_code'):
            # clear MFA session data
            for k in ('mfa_code', 'mfa_sent_at', 'last_resend_at'):
                session.pop(k, None)

            # reload user object and log them in
            with shelve.open('users.db', 'r') as db:
                stored = db.get(session['username'])
                user_uuid=stored['uuid']
            user = User(
                username= session['username'],
                email=    stored['email'],
                password= stored['password'],
                role=     stored['role'],
                status=   stored.get('status', 'Active')
            )
            login_user(user)
            session["uuid"] = user_uuid 
            return redirect(url_for('set_ip'))

        flash('Invalid authentication code', 'danger')

    # 5) Compute timers for template rendering
    expiry_timer = 0
    if (sent_iso := session.get('mfa_sent_at')):
        sent    = datetime.fromisoformat(sent_iso)
        elapsed = (datetime.now(timezone.utc) - sent).total_seconds()
        expiry_timer = int(max(0, 60 - elapsed))

    resend_timer = 0
    if (last_iso := session.get('last_resend_at')):
        last    = datetime.fromisoformat(last_iso)
        elapsed = (datetime.now(timezone.utc) - last).total_seconds()
        resend_timer = int(max(0, 5 - elapsed))

    return render_template(
        'mfa_verify.html',
        form=           form,
        expiry_timer=   expiry_timer,
        resend_timer=   resend_timer,
        hide_sidebar=True
    )
  
@app.route('/resend-mfa', methods=['POST'])
def resend_mfa():
    last_iso = session.get('last_resend_at')
    now      = datetime.now(timezone.utc)

    # only allow resend if ≥ 5 s have passed
    if not last_iso or (now - datetime.fromisoformat(last_iso)).total_seconds() >= 5:
        send_mfa_code()
        return ('', 204)

    retry_after = 5 - (now - datetime.fromisoformat(last_iso)).total_seconds()
    return (
        jsonify({
            'error': 'Too many requests',
            'retry_after': int(retry_after)
        }),
        429
    )


#mfa for forget_password
def send_reset_mfa_code():
    """
    Generate/store a 6-digit code for password-reset MFA and email it.
    """
    code = str(secrets.randbelow(900_000) + 100_000)
    print("The MFA Code: ", code)
    now = datetime.now(timezone.utc)

    session['reset_mfa_code']       = code
    session['reset_mfa_sent_at']    = now.isoformat()
    session['reset_last_resend_at'] = now.isoformat()
    current_app.logger.debug(f"[DEBUG] Reset MFA code is: {code}")

    text_body = (
        "I@FAD Password Reset Verification\n\n"
        f"Your 6-digit code is: {code}\n"
        "It expires in 60 seconds.\n\n"
        "If you didn’t request this, ignore this email."
    )
    html_body = render_template(
        'email/reset_mfa.html',
        code=code,
        expires=60,
        year=now.year
    )

    try:
        msg = flask_mail.Message(
            subject="I@FAD Password Reset Code",
            sender=current_app.config['MAIL_DEFAULT_SENDER'],
            recipients=[session.get('email')],
            body=text_body,
            html=html_body
        )
        mail.send(msg)
        return True
    except Exception as e:
        current_app.logger.error(f"[ERROR] Failed to send reset-MFA email: {e}")
        return False


@app.route('/forgetpassword', methods=['GET','POST'])
def forget_password():
    form = forgetpassword()
    if form.validate_on_submit():
        foremail = form.email.data

        # lookup user by email
        with shelve.open('users.db', 'r') as db:
            match = next(
                (u for u, data in db.items() if data.get('email') == foremail),
                None
            )

        if not match:
            flash('Email not found.', 'danger')
            return render_template('forget_password.html', form=form)

        # stash and send code (no flash here)
        session['username'] = match
        session['email']    = foremail

        if not send_reset_mfa_code():
            flash('Unable to send authentication code right now. Please try again later.', 'danger')
            return redirect(url_for('forget_password'))

        return redirect(url_for('mfa_verify2'))

    return render_template('forget_password.html', form=form, hide_sidebar=True)


@app.route('/mfa-verify2', methods=['GET','POST'])
def mfa_verify2():
    # ensure they came from forget_password
    if 'username' not in session or 'email' not in session:
        flash('Please submit your email first.', 'warning')
        return redirect(url_for('forget_password'))

    form     = MFAForm()
    now      = datetime.now(timezone.utc)
    sent_iso = session.get('reset_mfa_sent_at')

    # determine expired state
    expired = False
    if sent_iso:
        sent_at = datetime.fromisoformat(sent_iso)
        expired = (now - sent_at).total_seconds() > 60

    # ── GET: only flash on first send, silently re-send on expiry ──────────
    if request.method == 'GET':
        # first-ever GET: no code yet
        if not sent_iso:
            if not send_reset_mfa_code():
                flash('Unable to send authentication code right now. Please try again later.', 'danger')
                return redirect(url_for('forget_password'))
            flash('An authentication code has been sent to your email.', 'info')

        # subsequent GETs after expiry: re-send silently
        elif expired:
            send_reset_mfa_code()

    # ── POST: validate submitted code ───────────────────────────────────────
    if form.validate_on_submit():
        entered  = form.code.data
        sent_iso = session.get('reset_mfa_sent_at')
        now      = datetime.now(timezone.utc)

        if not sent_iso:
            flash('No code found. Please wait for the new one.', 'danger')
            return redirect(url_for('mfa_verify2'))

        sent_at = datetime.fromisoformat(sent_iso)
        if (now - sent_at).total_seconds() > 60:
            flash('Code expired. A new code has been sent.', 'danger')
            send_reset_mfa_code()
            return redirect(url_for('mfa_verify2'))

        if entered == session.get('reset_mfa_code'):
            # clear MFA bits
            for k in ('reset_mfa_code','reset_mfa_sent_at','reset_last_resend_at'):
                session.pop(k, None)

            flash('MFA verification successful. Please reset your password.', 'success')
            return redirect(url_for('reset_password'))

        flash('Invalid authentication code.', 'danger')

    # ── Compute timers for template ───────────────────────────────────────
    def _timer(key, limit):
        iso = session.get(key)
        if not iso:
            return 0
        elapsed = (datetime.now(timezone.utc) - datetime.fromisoformat(iso)).total_seconds()
        return int(max(0, limit - elapsed))

    expiry_timer = _timer('reset_mfa_sent_at', 60)
    resend_timer = _timer('reset_last_resend_at', 5)

    return render_template(
        'mfa_verify2.html',
        form=           form,
        expiry_timer=   expiry_timer,
        resend_timer=   resend_timer,
        hide_sidebar=True
    )


@app.route('/resend-mfa2', methods=['POST'])
def resend_mfa2():
    last_iso = session.get('reset_last_resend_at')
    now      = datetime.now(timezone.utc)

    if not last_iso or (now - datetime.fromisoformat(last_iso)).total_seconds() >= 5:
        send_reset_mfa_code()
        return ('', 204)

    retry_after = 5 - (now - datetime.fromisoformat(last_iso)).total_seconds()
    return (
        jsonify({
            'error': 'Too many requests',
            'retry_after': int(retry_after)
        }), 429
    )
# ── 4) Finally, reset the password ───────────────────────────────────────────

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    form = updatepasswordForm()

    # guard: must’ve done MFA
    if 'username' not in session or 'email' not in session:
        flash('You must verify your identity before resetting your password.', 'danger')
        return redirect(url_for('forget_password'))

    username = session['username']

    if form.validate_on_submit():
        p1 = form.password.data
        p2 = form.confirm_password.data

        if p1 != p2:
            flash('Passwords do not match.', 'danger')
        else:
            try:
                with shelve.open('users.db', 'w') as db:
                    if username in db:
                        user = db[username]
                        user['password'] = generate_password_hash(p1, method='pbkdf2:sha256')
                        db[username] = user
                        flash('Your password has been updated. Please log in with the new password.', 'success')
                        session.clear()
                        return redirect(url_for('login'))
                    else:
                        flash('User not found. Please start again.', 'danger')
                        return redirect(url_for('forget_password'))

            except Exception as e:
                current_app.logger.error(f"Error resetting password: {e}")
                flash('An error occurred. Please try again.', 'danger')

    return render_template('reset_password.html', form=form)

@app.route('/logout')
def logout():
    session.clear()

    # Clear the stored user email from shelve
    try:
        with shelve.open('settings.db', writeback=True) as db:
            if 'CurrentUserEmail' in db:
                del db['CurrentUserEmail']
    except Exception as e:
        print(f"Error clearing CurrentUserEmail: {e}")

    flash('Please log in to access.', 'info')
    return redirect(url_for('login'))


from flask import Flask, flash, render_template, request, redirect, session, url_for
import shelve, re

# Helper function to get settings from the database
def get_settings():
    with shelve.open('settings.db', 'c') as db:
        settings = db.get('settings', {'interval': 2000, 'threshold': 5})  # Default interval is 2000ms and default threshold is 5
        app.logger.debug(f"Settings read from database: {settings}")
    return settings

def update_settings(new_settings):
    with shelve.open('settings.db', 'c') as db:
        db['settings'] = new_settings
        app.logger.debug(f"Settings updated in database: {new_settings}")

from scapy.all import *

def get_next_ip_id():
    with shelve.open('settings.db', 'c') as db:
        last_id = db.get('last_ip_id', 7500)
        db['last_ip_id'] = last_id + 1
        return last_id
# 
def get_interface_by_ip(target_ip):
    for iface_name, iface_addrs in psutil.net_if_addrs().items():
        for addr in iface_addrs:
            if addr.family == socket.AF_INET and addr.address == target_ip:
                print(f"Found interface {iface_name} with IP {target_ip}")
                return iface_name
    raise RuntimeError(f"No interface found with IP {target_ip}")

def send_tcp_packet(encoded_byte, source_port, server_isn, server_ack, source_ip, dest_ip):
    destination_ip = dest_ip
    source_ip = source_ip
    destination_port = 50000

    payload = bytes.fromhex(encoded_byte)
    flags = "PA"  # PSH + ACK flags
    ip_id = get_next_ip_id()

    # Detect the correct network interface
    try:
        interface_name = get_interface_by_ip(source_ip)
        print(f"Using interface: {interface_name}")
    except RuntimeError as e:
        print(f"[Error] {e}")
        return

    # Create IP and TCP headers without manually setting checksum
    ip_packet = IP(src=source_ip, dst=destination_ip, id=ip_id, flags="DF", ttl=128)

    # Let Scapy calculate the checksum automatically
    tcp_packet = TCP(
        sport=source_port,
        dport=destination_port,
        seq=server_isn,
        ack=server_ack,
        flags=flags,
        window=64233
    )

    packet = ip_packet / tcp_packet / payload
    print(f"Sent TCP packet with payload: {encoded_byte}")
    send(packet)

    packet_counter = 0

    while True:
        packet_counter += 1
        print(f"Sniff attempt {packet_counter}")  # <-- Add this
        sniffpacket = sniff(iface=interface_name,
                            filter=f"tcp and src host {destination_ip} and port {destination_port}",
                            count=1, timeout=0.5) # Adjust the Ethernet port name to correspond with the connected interface.

        if not sniffpacket:
            print(f"No packet received. Attempt: {packet_counter}")  # <-- Add this
            if packet_counter == 2:
                print("Timeout reached, no packet captured.")
                return False
            continue

        pkt = sniffpacket[0]
        tcp_flags = pkt[TCP].flags
        tcp_payload = bytes(pkt[TCP].payload).rstrip(b'\x00')
        payload_len = len(tcp_payload)
        print(f"Packet captured: {payload_len} bytes | Flags: {tcp_flags}")

        if tcp_flags in ["PA", "A", "FPA", "FA"]:
            stored_ack = sniffpacket[0][TCP].ack
            stored_seq = sniffpacket[0][TCP].seq

            # Send ACK with proper sequence and acknowledgment numbers
            ip_id = get_next_ip_id()
            ack_packet = IP(src=source_ip, dst=destination_ip, id=ip_id) / TCP(
                sport=source_port,
                dport=destination_port,
                seq=stored_ack,
                ack=stored_seq + payload_len,
                flags="A",
                window=64233
            )
            send(ack_packet)
            with shelve.open('settings.db', 'c') as db:
                db['syn_ack_seq'] = stored_ack
                db['syn_ack_ack'] = stored_seq + payload_len

            print("Found correct packet and sent ACK")
            return True
        else:
            continue

# Function to perform the 3-way TCP handshake
def send_syn_packet(client_ip, server_ip, source_port, destination_port):
    try:
        try:
            interface_name = get_interface_by_ip(client_ip)
            print(f"Using interface: {interface_name}")
        except RuntimeError as e:
            print(f"[Error] {e}")
            return
        # Step 1: Generate a random ISN
        isn = random.randint(0, 2**32 - 1)
        ip_id = get_next_ip_id()
        # Step 2: Create and send the SYN packet
        ip = IP(src=client_ip, dst=server_ip,id=ip_id)
        tcp_options = [
            ('MSS', 1460),          # Maximum Segment Size
            ('NOP', None),          # No-Operation (NOP)
            ('WScale', 8),          # Window Scale
            ('NOP', None),          # No-Operation (NOP)
            ('NOP', None),          # No-Operation (NOP)
            ('SAckOK', b'')         # SACK Permitted
        ]
        syn = TCP(
            sport=source_port,
            dport=destination_port,
            flags="S",
            seq=isn,  # Set the ISN
            window=64240,
            options=tcp_options,
            chksum = None
        )

        syn_ack = sr1(ip / syn, timeout=1)

        if syn_ack is None or syn_ack[TCP].flags != "SA":
            print("Did not receive SYN-ACK")
            return None, None
        ip_id = get_next_ip_id()
        ip = IP(src=client_ip, dst=server_ip,id=ip_id)
        # Step 3: Send the ACK to complete the handshake
        ack = TCP(
            sport=source_port,
            dport=destination_port,
            flags="A",
            seq=syn_ack.ack,  # Increment by 1 as per TCP handshake
            ack=syn_ack.seq + 1,  # Acknowledge the server's SYN-ACK
            window=64240,
            chksum = None
        )
        send(ip / ack)
        print("Handshake completed")
        time.sleep(0.2)

        while True:

            psh_ack = sniff(iface=interface_name, filter=f"tcp and src host {server_ip} and port {destination_port}", count=1, timeout=1) # Adjust the Ethernet port name to correspond with the connected interface.

            if not psh_ack:
                # print("no packet")
                continue

            psh_ack = psh_ack[0][TCP]

            if  psh_ack.flags == "PA" : # PSH-ACK received

                payload_data = bytes(psh_ack.payload).rstrip(b'\x00')

                payload_length = len(payload_data)

                print("Keep-Alive PSH-ACK received")

                # Send Keep-Alive ACK
                ip_id = get_next_ip_id()
                ip = IP(src=client_ip, dst=server_ip, id=ip_id)
                keep_alive_ack = TCP(

                    sport=source_port,

                    dport=destination_port,

                    flags="A",

                    seq=psh_ack.ack,  # Use the ACK number from the PSH-ACK

                    ack=psh_ack.seq + payload_length,  # Acknowledge the Keep-Alive payload

                    window=64236,
                    chksum = None
                )
                print(f"actual seq = {psh_ack.ack}")

                send(ip / keep_alive_ack)
                with shelve.open('settings.db', 'c') as db:
                    db['syn_ack_seq'] = psh_ack.ack
                    db['syn_ack_ack'] = psh_ack.seq + payload_length
                print("Acknowledged server's PSH, ACK")
                print(f"ISN{syn_ack.seq}")

                print("Keep-Alive ACK sent")
                time.sleep(1)
            else:
                print("Unexpected packet received")
            # Save the SYN-ACK sequence number for further use

    except Exception as e:
        print(f"Error: {e}")

def send_RST_packet(source_port, server_isn, server_ack, source_ip, destination_ip):
    destination_port = 50000
    ip = IP(src=source_ip, dst=destination_ip)

    # Construct the TCP header
    tcp = TCP(
        sport=source_port,  # Source port
        dport=destination_port,  # Destination port
        flags="RA",  # RST and ACK flags
        seq=server_ack,  # Sequence number
        ack=server_isn+4,  # Acknowledgment number
        window=0,  # Window size
    )

    # Combine the headers into a single packet
    packet = ip / tcp

    # Send the packet
    send(packet, verbose=0)

def start_send_manual_feed(source_port, server_isn, server_ack, source_ip, destination_ip):
    print(f"[DEBUG] --> start_send_manual_feed() triggered at {datetime.now().strftime('%H:%M:%S')}")
    encoded_data = "ccdda10100010001a448"
    
    max_retries = 5
    for attempt in range(1, max_retries + 1):
        success = send_tcp_packet(
            encoded_byte=encoded_data,
            source_port=source_port,
            server_isn=server_isn,
            server_ack=server_ack,
            source_ip=source_ip,
            dest_ip=destination_ip
        )
        if success:
            print(f"[DEBUG] Start feed packet sent successfully on attempt {attempt}")
            return True
        else:
            print(f"[WARNING] Start Manual feed packet not sent (attempt {attempt})")
    else:
        print("[ERROR] Failed to send Start Manual feed packet after 5 attempts")
        return False

def stop_send_manual_feed(source_port, server_isn, server_ack, source_ip, destination_ip):
    print(f"[DEBUG] --> stop_send_manual_feed() triggered at {datetime.now().strftime('%H:%M:%S')}")
    encoded_data = "ccdda10100000001a346"
    
    max_retries = 5
    for attempt in range(1, max_retries + 1):
        success = send_tcp_packet(
            encoded_byte=encoded_data,
            source_port=source_port,
            server_isn=server_isn,
            server_ack=server_ack,
            source_ip=source_ip,
            dest_ip=destination_ip
        )
        if success:
            print(f"[DEBUG] Stop feed packet sent successfully on attempt {attempt}")
            return True
        else:
            print(f"[WARNING] Stop Manual feed packet not sent (attempt {attempt})")
    else:
        print("[ERROR] Failed to send Stop Manual feed packet after 5 attempts")
        return False

# Route to update the interval setting
@app.route('/update_interval', methods=['POST'])
def update_interval():
    try:
        minutes = request.form.get('interval_minutes', 0)  # Retrieve from form
        seconds = request.form.get('interval_seconds', 0)

        app.logger.debug(f"Received minutes: {minutes}, seconds: {seconds}")

        minutes = int(minutes) if minutes else 0
        seconds = int(seconds) if seconds else 0

        if minutes < 0 or seconds < 0:
            raise ValueError("Minutes and seconds should be non-negative integers.")

        total_seconds = (minutes * 60) + seconds

        # Update settings
        settings = get_settings()
        settings['interval_seconds'] = total_seconds
        update_settings(settings)

        return redirect(url_for('dashboard'))  # Redirect after updating

    except ValueError as e:
        flash('Invalid input! Minutes and seconds must be non-negative integers.', 'danger')
        return redirect(url_for('settings'))  # Redirect to settings page

    except Exception as e:
        app.logger.error(f"An error occurred while updating interval: {str(e)}")
        flash('An unexpected error occurred.', 'danger')
        return redirect(url_for('settings'))

# Route to retrieve the current interval setting
@app.route('/get_interval', methods=['GET'])
def get_interval():
    try:
        settings = get_settings()
        app.logger.debug(f"Current interval retrieved: {settings['interval']}")
        return jsonify({'interval': settings['interval']}), 200
    except Exception as e:
        app.logger.error(f"An error occurred while retrieving interval: {str(e)}")
        return jsonify({'error': 'An error occurred while retrieving interval.'}), 500

@app.route('/update_threshold', methods=['POST'])
def update_threshold():
    try:
        threshold = request.json.get('threshold')
        app.logger.debug(f"Received threshold: {threshold}")  # Debug log

        # Validate threshold (should be a positive integer)
        if threshold is not None:
            threshold = int(threshold)
            if threshold <= 0:
                raise ValueError("Threshold should be a positive integer.")

            # Update threshold in settings
            settings = get_settings()
            settings['threshold'] = threshold
            update_settings(settings)
            app.logger.debug(f"Threshold updated successfully to {threshold}")
            return jsonify({'message': 'Threshold updated successfully'}), 200
        else:
            return jsonify({'error': 'Threshold value not provided'}), 400

    except ValueError as e:
        app.logger.error(f"Invalid threshold value: {e}")
        return jsonify({'error': 'Invalid threshold value. Must be a positive integer.'}), 400

    except Exception as e:
        app.logger.error(f"An error occurred while updating threshold: {str(e)}")
        return jsonify({'error': 'An error occurred while updating threshold.'}), 500

@app.route('/get_threshold', methods=['GET'])
def get_threshold():
    try:
        settings = get_settings()
        app.logger.debug(f"Current threshold retrieved: {settings['threshold']}")
        return jsonify({'threshold': settings['threshold']}), 200
    except Exception as e:
        app.logger.error(f"An error occurred while retrieving threshold: {str(e)}")
        return jsonify({'error': 'An error occurred while retrieving threshold.'}), 500
    
def is_valid(val):
    try:
        return float(val) >= 0
    except:
        return False

@app.route('/pellet_data')
def get_pellet_data():
    # Define test data
    pellet_data = {
        '26 Jul 2025': {'Morning Feed 1': 250, 'Morning Feed 2': "", 'Evening Feed 1': 200, 'Evening Feed 2': 250, 'Total': 750},
        '27 Jul 2025': {'Morning Feed 1': 200, 'Morning Feed 2': 250, 'Evening Feed 1': 200, 'Evening Feed 2': 200, 'Total': 850},
        '28 Jul 2025': {'Morning Feed 1': 100, 'Morning Feed 2': 0, 'Evening Feed 1': 200, 'Evening Feed 2': 250, 'Total': 550},
        '29 Jul 2025': {'Morning Feed 1': 120, 'Morning Feed 2': 180, 'Evening Feed 1': 130, 'Evening Feed 2': 170, 'Total': 600},
        '30 Jul 2025': {'Morning Feed 1': 220, 'Morning Feed 2': "N/A", 'Evening Feed 1': 180, 'Evening Feed 2': 230, 'Total': 630},
        '31 Jul 2025': {'Morning Feed 1': 140, 'Morning Feed 2': 190, 'Evening Feed 1': 160, 'Evening Feed 2': 210, 'Total': 700},
        '1 Aug 2025': {'Morning Feed 1': 125, 'Morning Feed 2': 175, 'Evening Feed 1': 150, 'Evening Feed 2': 200, 'Total': 650},
    }
    # Check if the database exists, and if not, create and populate it
    db_path = 'mock_chart_data.db'
    if not os.path.exists(db_path + '.db'):  # Shelve creates additional files with ".db"
        with shelve.open(db_path, 'c') as db:
            for date, feeds in pellet_data.items():
                db[datetime.strptime(date, "%d %b %Y").strftime("%Y-%m-%d")] = feeds
            db.sync()
        print("Database created and initialized.")

    # Generate the last 7 days (including today)
    current_day = datetime.today().date()
    last_7_days = [(current_day - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(6, -1, -1)]

    # Initialize lists to hold pellet counts
    morning_1 = []
    morning_2 = []
    evening_1 = []
    evening_2 = []
    total = []
    session_counts = []

    with shelve.open(db_path, 'r') as db:
        for day in last_7_days:
            data = db.get(day, {})
            morning_feed_1 = data.get('Morning Feed 1', 'N/A')
            morning_feed_2 = data.get('Morning Feed 2', 'N/A')
            evening_feed_1 = data.get('Evening Feed 1', 'N/A')
            evenin_feed_2 = data.get('Evening Feed 2', 'N/A')

            morning_1.append(morning_feed_1)
            morning_2.append(morning_feed_2)
            evening_1.append(evening_feed_1)
            evening_2.append(evenin_feed_2)

            valid_values = [v for v in [morning_feed_1, morning_feed_2, evening_feed_1, evenin_feed_2] if is_valid(v)]
            valid_total = sum(float(v) for v in valid_values)
            total.append(valid_total)
            session_counts.append(len(valid_values))

    return jsonify({
        'labels': [datetime.strptime(day, "%Y-%m-%d").strftime("%d %b") for day in last_7_days],
        'morning_1': morning_1,
        'morning_2': morning_2,
        'evening_1': evening_1,
        'evening_2': evening_2,
        'total': total,
        'session_counts': session_counts
    })

@app.route('/dashboard', methods=['GET'])
@login_required
def dashboard():
    user_settings = {}
    latest_count = 0
    pellet_data = {}
    try:
        with shelve_lock:
            with shelve.open("users.db", "r") as user_db:
                for user_data in user_db.values():
                    if user_data.get("uuid") == session.get("uuid"):
                        user_settings = {
                            "first_morning_timer": user_data.get("user_morning_feed_1", ""),
                            "first_evening_timer": user_data.get("user_evening_feed_1", ""),
                            "second_morning_timer": user_data.get("user_morning_feed_2", ""),
                            "second_evening_timer": user_data.get("user_evening_feed_2", ""),
                            "minutes": user_data.get("user_minutes", 0),
                            "interval_seconds": user_data.get("user_interval_seconds", 0),
                            "pellets": user_data.get("user_pellets", 0),
                            "pellet_size": user_data.get("pellet_size", 1),
                            "pellets_per_second": user_data.get("pellets_per_second", 1),
                            "feeding_threshold": user_data.get("feeding_threshold", 0),
                        }
                        break


        with shelve.open('currentcount.db', 'c') as db2:
            latest_count = db2.get('object_count', 0)

        print("Fetching pellet data")
        response = get_pellet_data()
        pellet_data = response.json

    except Exception as e:
        print(f"[ERROR] Failed to load dashboard data: {e}")

    return render_template(
        'dashboard.html', user_settings=user_settings, latest_count=latest_count, pellet_labels=pellet_data.get('labels', []), first_feed_left=pellet_data.get('first_feed_left', []),
        second_feed_left=pellet_data.get('second_feed_left', []), total_feed_count=pellet_data.get('total_feed_count', []), checking_interval=user_settings.get("interval_seconds", 10))

@app.route('/camera_view',methods=['GET','POST'])
@login_required
def camera_view():
    return render_template('camera_view.html')

@app.route('/export_data', methods=['POST'])
def export_data():

    data = request.get_json()
    print(f"Received data for export: {data}")
    labels = data.get('labels', [])
    morning_1 = data.get('morning_1', [])
    morning_2 = data.get('morning_2', [])
    evening_1 = data.get('evening_1', [])
    evening_2 = data.get('evening_2', [])
    total = data.get('total', [])
    feeding_rate = data.get('feeding_rate', [])

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    todayDate = datetime.today().strftime("%Y-%m-%d")
    seven_days_ago = (datetime.today() - timedelta(days=6)).strftime("%Y-%m-%d")
    sheet.title = f"Feeding Summary ({seven_days_ago} - {todayDate})"

    # Title Row
    sheet.merge_cells("A1:G1")
    sheet["A1"] = f"Past 7 Days' Feeding Summary ({seven_days_ago} - {todayDate})"
    sheet["A1"].font = Font(bold=True, size=14)

    # Header Row
    headers = ["Date", "Morning Feed 1", "Morning Feed 2", "Evening Feed 1", "Evening Feed 2", "Total Feed Amount", "Feeding Rate"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    # Populate data
    for i in range(len(labels)):
        row = i + 3
        morning_feed_1 = morning_1[i]
        morning_feed_2 = morning_2[i]
        evening_feed_1 = evening_1[i]
        evenin_feed_2 = evening_2[i]
        total_feed = total[i] if isinstance(total[i], (int, float)) else 0
        day_feeding_rate = feeding_rate[i]
        

        sheet.cell(row=row, column=1, value=labels[i])
        sheet.cell(row=row, column=2, value=morning_feed_1 if morning_feed_1 != "" else "N/A")
        sheet.cell(row=row, column=3, value=morning_feed_2 if morning_feed_2 != "" else "N/A")
        sheet.cell(row=row, column=4, value=evening_feed_1 if evening_feed_1 != "" else "N/A")
        sheet.cell(row=row, column=5, value=evenin_feed_2 if evenin_feed_2 != "" else "N/A")
        sheet.cell(row=row, column=6, value=total_feed)
        sheet.cell(row=row, column=7, value=f"{day_feeding_rate}%")

    # Column width adjustment
    column_widths = [20, 20, 20, 20, 20, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    file_path = 'feed_summary_data.xlsx'
    workbook.save(file_path)

    return send_file(file_path, as_attachment=True, download_name='feed_summary_data.xlsx')

import re
latest_valid_settings = None
latest_set_ip_settings = None
active_feeding_user_lock = threading.Lock()
active_feeding_user = {
    "uuid": None,
    "email": None,
    "config": None,
}

@app.route('/update', methods=['GET', 'POST'])
@login_required
def update_setting():
    setting = configurationForm(request.form)
    mode = request.args.get("mode", "auto")
    try:
        if request.method == 'POST':
            if request.is_json:
                data = request.get_json()

                # MANUAL FEED BLOCK
                if mode == "manual":
                    manual_form = data.get("manual_form")
                    manual_feed_action = data.get("manual_feed_action")
                    if manual_form and manual_feed_action in ["start", "stop"]:
                        try:
                            with shelve.open("settings.db", 'c') as db, shelve.open("IP.db", 'r') as ip_db:
                                config = {
                                    'Port': db.get('Port'),
                                    'syn_ack_seq': db.get('syn_ack_seq'),
                                    'syn_ack_ack': db.get('syn_ack_ack'),
                                    'source': ip_db.get("IP", {}).get("source"),
                                    'destination': ip_db.get("IP", {}).get("destination")
                                }

                            for key, value in config.items():
                                if not value:
                                    location = "IP.db" if key in ['source', 'destination'] else "settings"
                                    return jsonify({"status": "error", "message": f"Missing '{key}' in {location}."}), 400

                            port = config['Port']
                            server_isn = config['syn_ack_seq']
                            server_ack = config['syn_ack_ack']
                            source_ip = config['source']
                            destination_ip = config['destination']

                            if manual_feed_action == "start":
                                success = start_send_manual_feed(port, server_isn, server_ack, source_ip, destination_ip)
                                if success:
                                    return jsonify({"status": "success", "message": "Manual feeding started."})
                                else:
                                    return jsonify({"status": "error", "message": "Failed to start manual feeding."}), 500
                            else:
                                success = stop_send_manual_feed(port, server_isn, server_ack, source_ip, destination_ip)
                                if success:
                                    return jsonify({"status": "success", "message": "Manual feeding stopped."})
                                else:
                                    return jsonify({"status": "error", "message": "Failed to stop manual feeding."}), 500
                        except Exception as e:
                            return jsonify({"status": "error", "message": f"Manual feed error: {e}"}), 500

                    return jsonify({"status": "error", "message": "Invalid manual feed request."}), 400

                # AUTO FEED BLOCK
                elif mode == "auto":
                    required_fields = [
                        "user_morning_feed_1", "user_evening_feed_1", "user_morning_feed_2","user_evening_feed_2", "user_minutes", "user_interval_seconds", "user_pellets",
                        "user_feeding_threshold", "user_pellet_size", "user_pellets_per_second"]
                    if not all(k in data for k in required_fields):
                        return jsonify({"status": "error", "message": "Missing required fields."}), 400
                    
                    print(f"[DEBUG] Received required fields: {required_fields}")
                    
                    try:
                        # Save all form inputs
                        config_values = {
                            "user_morning_feed_1": data['user_morning_feed_1'],
                            "user_morning_feed_2": data['user_morning_feed_2'] or "N/A",
                            "user_evening_feed_1": data['user_evening_feed_1'],
                            "user_evening_feed_2": data['user_evening_feed_2'] or "N/A",
                            "user_minutes": int(data['user_minutes']),
                            "user_interval_seconds": int(data['user_interval_seconds'] or 1),
                            "user_pellets": int(data['user_pellets'] or 0),
                            "user_feeding_threshold": int(data['user_feeding_threshold'] or 0),
                            "user_pellet_size": int(data['user_pellet_size'] or 1),
                            "user_pellets_per_second": int(data['user_pellets_per_second'] or 1)
                        }
                        print(f"[DEBUG] Config values to save: {config_values}")

                        # Save to user DB
                        with shelve.open("users.db", 'w') as user_db:
                            for username, user_data in user_db.items():
                                if user_data.get("uuid") == session.get("uuid"):
                                    user_data["user_morning_feed_1"] = data['user_morning_feed_1'] or "N/A"
                                    user_data["user_morning_feed_2"] = data['user_morning_feed_2'] or "N/A"
                                    user_data["user_evening_feed_1"] = data['user_evening_feed_1'] or "N/A"
                                    user_data["user_evening_feed_2"] = data['user_evening_feed_2'] or "N/A"

                                    user_data["user_minutes"] = int(data['user_minutes'] or 0)
                                    user_data["user_interval_seconds"] = int(data['user_interval_seconds'] or 0)
                                    user_data["user_pellets"] = int(data['user_pellets'] or 0)
                                    user_data["user_feeding_threshold"] = int(data['user_feeding_threshold'] or 0)  # no user_ prefix per your initial data
                                    user_data["user_pellet_size"] = int(data['user_pellet_size'] or 1)
                                    user_data["user_pellets_per_second"] = int(data['user_pellets_per_second'] or 1)
                                    print(f"[USER FEED SETTINGS] Updated for {username}")
                                    print(f"[USER FEED SETTINGS] Data: {user_data}")
                                    user_db[username] = user_data
                                    break

                        # Activate in-memory user feeding config
                        set_active_feeding_user(session.get("uuid"), session.get('email'), config_values)

                        # Check system dependencies
                        with shelve.open("settings.db", 'r') as db, shelve.open("IP.db", 'r') as ip_db:
                            config = {
                                'Port': db.get('Port'),
                                'syn_ack_seq': db.get('syn_ack_seq'),
                                'syn_ack_ack': db.get('syn_ack_ack'),
                                'source': ip_db.get("IP", {}).get("source"),
                                'destination': ip_db.get("IP", {}).get("destination")
                            }
                        for key, value in config.items():
                            if not value:
                                location = 'IP.db' if key in ['source', 'destination'] else 'settings'
                                return jsonify({"status": "error", "message": f"Missing '{key}' in {location}."}), 400

                        reschedule_feeding_alerts()
                        return jsonify({"status": "success", "message": "Feeding schedule updated."})

                    except Exception as e:
                        return jsonify({"status": "error", "message": f"Auto feed setup failed: {e}"}), 500

                return jsonify({"status": "error", "message": "Invalid mode."}), 400

            # Fallback for form-based submission
            return render_template('settings.html', form=setting, mode=mode)

        # GET method – load current settings
        with shelve.open("users.db", 'r') as db:
            for user_data in db.values():
                if user_data.get("uuid") == session.get("uuid"):
                    print(user_data)
                    setting.morning_feed_1.data = user_data.get("user_morning_feed_1", "")
                    setting.morning_feed_2.data = user_data.get("user_morning_feed_2", "")
                    setting.evening_feed_1.data = user_data.get("user_evening_feed_1", "")
                    setting.evening_feed_2.data = user_data.get("user_evening_feed_2", "")
                    setting.minutes.data = user_data.get("user_minutes", 0)
                    setting.interval_seconds.data = user_data.get("user_interval_seconds", 0)
                    setting.pellets.data = user_data.get("user_pellets", 0)
                    setting.feeding_threshold.data = user_data.get("user_feeding_threshold", 0)
                    setting.pellet_size.data = user_data.get("user_pellet_size", 0)
                    setting.pellets_per_second.data = user_data.get("user_pellets_per_second", 0)
                    break

        return render_template('settings.html', form=setting, mode=mode)

    except Exception as e:
        return jsonify({"status": "error", "message": f"Unexpected server error: {e}"}), 500
    
def set_active_feeding_user(user_uuid, user_email, feeding_config):
    with active_feeding_user_lock:
        active_feeding_user["uuid"] = user_uuid
        active_feeding_user["email"] = user_email
        active_feeding_user["config"] = feeding_config

def get_active_feeding_user():
    with active_feeding_user_lock:
        return active_feeding_user.copy()

def send_feeding_complete_email(user_email, feed_time_str, session_type):
    with app.app_context():
        try:
            with shelve_lock:
                with shelve.open('settings.db', 'r') as db:
                    Email_dict = db.get('Email_Data', {})
                    email_info = Email_dict.get('Email_Info')
                    support_email = email_info.get_sender_email() if email_info else "iatfadteam@gmail.com"

            msg = flask_mail.Message(
                subject="Feeding Complete",
                recipients=[user_email]
            )

            msg.body = f"The {session_type} feeding session at {feed_time_str} has been completed."


            msg.html = render_template(
                'email/feeding_complete.html',
                feed_time=feed_time_str,
                session_type=session_type.capitalize(),
                year=datetime.now().year,
                support_email=support_email
            )

            mail.send(msg)
            print(f"Email sent to {user_email} for {session_type} feeding at {feed_time_str}.")

        except Exception as e:
            print(f"Error sending email: {e}")

def reschedule_feeding_alerts():
    try:
        active_user = get_active_feeding_user()
        user_uuid = active_user.get("uuid")
        user_email = active_user.get("email", "iatfadteam@gmail.com")
        config = active_user.get("config")

        print(f"[RESCHEDULE] Rescheduling alerts for user UUID: {user_uuid}")
        print(f"[RESCHEDULE] Config: {config}")

        if not user_uuid or not config:
            print("[RESCHEDULE] No active feeding user. Skipping alert scheduling.")
            return

        timers = {
            "morning_feed_1": config.get("user_morning_feed_1", ""),
            "morning_feed_2": config.get("user_morning_feed_2", ""),
            "evening_feed_1": config.get("user_evening_feed_1", ""),
            "evening_feed_2": config.get("user_evening_feed_2", "")
        }

        feeding_duration_sec = int(config.get("user_minutes", 0)) * 60
        if feeding_duration_sec <= 0:
            print(f"[RESCHEDULE] Duration is zero or invalid for {user_email}")
            return

        for feed_key, time_str in timers.items():
            feed_id = f"{feed_key}_alert"

            if not time_str or not time_str.isdigit() or len(time_str) != 4:
                if scheduler.get_job(feed_id):
                    scheduler.remove_job(feed_id)
                    print(f"[RESCHEDULE] Removed existing job for {feed_key}")
                print(f"[RESCHEDULE] Skipping {feed_key} (empty or invalid)")
                continue

            hour = int(time_str[:2])
            minute = int(time_str[2:])
            feed_end = (datetime(2000, 1, 1, hour, minute) + timedelta(seconds=feeding_duration_sec))

            session_label = "Morning" if "morning" in feed_key else "Evening"

            if scheduler.get_job(feed_id):
                scheduler.remove_job(feed_id)

            scheduler.add_job(
                func=send_feeding_complete_email,
                trigger='cron',
                hour=feed_end.hour,
                minute=feed_end.minute,
                args=[user_email, time_str, session_label],
                id=feed_id,
                replace_existing=True,
                misfire_grace_time=3600
            )

            print(f"[RESCHEDULE] Scheduled {feed_key} at {feed_end.strftime('%H:%M')} for {user_email}")

    except Exception as e:
        print(f"[RESCHEDULE ERROR] {e}")

@app.route('/update/email', methods=['GET', 'POST'])
@login_required
@role_required("Admin")
def update_email_settings():
    setting = emailForm(request.form)

    if request.method == 'POST' and setting.validate():
        try:
            with shelve_lock:
                with shelve.open('settings.db', 'w') as db:
                    Email_dict = db.get('Email_Data', {})
                    j = Email_dict.get('Email_Info')

                    if j:
                        j.set_sender_email(setting.sender_email.data)
                        j.set_recipient_email(setting.recipient_email.data)
                        j.set_APPPassword(setting.App_password.data)
                        j.set_days(setting.days.data)
                        Email_dict['Email_Info'] = j

                    db['Email_Data'] = Email_dict
                    db['Confidence_Rate'] = setting.confidence.data
                    print("Confidence rate set to", setting.confidence.data)

            return jsonify(success=True, message="Settings updated successfully"), 200
        except Exception as e:
            print("[ERROR] Failed to save email settings:", e)
            return jsonify(success=False, message="Failed to save settings"), 500

    else:
        with shelve_lock:
            with shelve.open('settings.db', 'r') as db:
                Email_dict = db.get('Email_Data', {})
                j = Email_dict.get('Email_Info')

                setting.sender_email.data = j.get_sender_email()
                setting.recipient_email.data = j.get_recipient_email()
                setting.App_password.data = j.get_APPPassword()
                setting.days.data = j.get_days()

                setting.confidence.data = db.get('Confidence_Rate', 60)

        return render_template('email_settings.html', form=setting)

@app.route('/clear_video_feed_access', methods=['GET'])
def clear_video_feed_access():
    with shelve.open('settings.db', 'w') as db:
        db['Generate_Status'] = False
    stop_camera_threads()
    return '', 204  # No Content (sendBeacon doesn't expect a response body)

@app.route('/video_feed')
def video_feed():
    with shelve.open('settings.db', 'w') as db:
        db['Generate_Status'] = True
    start_camera_threads()
    try:
        return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')
    except Exception as e:
        print(f"Error: {e}")
        return "Error generating video feed"

def send_feedback_notification(name, user_email, message, sent_time=None):
    """
    Send a notification email to the I@FAD team with the user's feedback,
    timestamped in Singapore time.
    """
    sg_tz = ZoneInfo("Asia/Singapore")
    now = sent_time or datetime.now(sg_tz)
    
    try:
        db = shelve.open('settings.db', 'r')
        Email_dict = db.get('Email_Data', {})
        recipient_email = "iatfadteam@gmail.com"  # Default fallback

        if 'Email_Info' in Email_dict:
            email_info = Email_dict['Email_Info']
            if hasattr(email_info, 'get_recipient_email'):
                configured_email = email_info.get_recipient_email()
                if configured_email:
                    recipient_email = configured_email
        db.close()
    except Exception as e:
        recipient_email = "iatfadteam@gmail.com"
        app.logger.warning(f"Could not retrieve recipient email, using fallback: {e}")

    # Plain-text body
    text_body = (
        "I@FAD New Feedback Received\n\n"
        f"From: {name} <{user_email}>\n\n"
        "Message:\n"
        f"{message}\n\n"
        f"Sent at {now.isoformat()}"
    )

    # HTML body
    html_body = render_template(
        'email/feedback.html',
        name=name,
        email=user_email,
        message=message,
        sent_at=now.strftime('%Y-%m-%d %H:%M'),
        year=now.year
    )

    msg = flask_mail.Message(
        subject="I@FAD – New Feedback",
        sender=recipient_email,
        recipients=[recipient_email],
        body=text_body,
        html=html_body
    )
    mail.send(msg)

def send_feedback_confirmation(user_name, user_email, message, sent_time=None):
    """
    Send a copy of the user’s feedback back to them as confirmation.
    """
    sg_tz = ZoneInfo("Asia/Singapore")
    now = sent_time or datetime.now(sg_tz)
    
    try:
        db = shelve.open('settings.db', 'r')
        Email_dict = db.get('Email_Data', {})
        recipient_email = "iatfadteam@gmail.com"  # Default fallback

        if 'Email_Info' in Email_dict:
            email_info = Email_dict['Email_Info']
            if hasattr(email_info, 'get_recipient_email'):
                configured_email = email_info.get_recipient_email()
                if configured_email:
                    recipient_email = configured_email
        db.close()
    except Exception as e:
        recipient_email = "iatfadteam@gmail.com"
        app.logger.warning(f"Could not retrieve recipient email, using fallback: {e}")

    # Plain-text fallback
    text_body = (
        "I@FAD Feedback Confirmation\n\n"
        f"Hello {user_name},\n\n"
        "Thanks for your feedback! Here's what we received:\n\n"
        f"{message}\n\n"
        f"Submitted at {now.strftime('%Y-%m-%d %H:%M')}\n\n"
        "We'll review it and get back to you if needed."
    )

    # HTML version
    html_body = render_template(
        'email/feedback_confirm.html',
        name=user_name,
        message=message,
        sent_at=now.strftime('%Y-%m-%d %H:%M'),
        year=now.year
    )

    msg = flask_mail.Message(
        subject="I@FAD – We Received Your Feedback",
        sender=recipient_email,
        recipients=[user_email],
        body=text_body,
        html=html_body
    )
    mail.send(msg)

@app.route('/feedback', methods=['GET', 'POST'])
@login_required
def feedback():
    form = FeedbackForm()
    if form.validate_on_submit():
        # 1) store feedback
        store_Feedback.add(
            user_name=current_user.username,
            user_email=current_user.email,
            message=form.message.data
        )

        # 2) notify the team
        send_feedback_notification(
            name=current_user.username,
            user_email=current_user.email,
            message=form.message.data
        )

        # 3) send confirmation to the user
        send_feedback_confirmation(
            user_name=current_user.username,
            user_email=current_user.email,
            message=form.message.data
        )

        flash('Your feedback has been sent successfully—and a copy has been emailed to you.', 'success')
        return redirect(url_for('feedback'))

    return render_template('feedback.html', form=form)

@app.route('/admin/feedbacks')
@login_required
@role_required('Admin')
def admin_feedbacks():
    # grab the raw list
    all_feedbacks = store_Feedback.list_all()

    # pull the search term from ?q=…
    q = request.args.get('q', '').strip().lower()

    if q:
        def matches(fb):
            return (
                q in fb.get_user_email().lower() or
                q in fb.get_user_name().lower() or
                q in fb.get_message().lower() or
                q in fb.get_submitted_at().strftime('%Y-%m-%d %H:%M').lower()
            )
        feedbacks = [fb for fb in all_feedbacks if matches(fb)]
    else:
        feedbacks = all_feedbacks

    feedbacks = sorted(
        feedbacks,
        key=lambda fb: fb.get_submitted_at(),
        reverse=True
    )

    form = DeleteForm()
    return render_template('admin_feedback.html',
                           feedbacks=feedbacks,
                           form=form,
                           q=q)

@app.route('/admin/feedbacks/delete/<int:fb_id>', methods=['POST'])
@login_required
@role_required('Admin')
def delete_feedback(fb_id):
    if store_Feedback.delete(fb_id):
        return jsonify({'success': True, 'message': f'Feedback deleted successfully.'}), 200
    else:
        return jsonify({'success': False, 'message': 'Could not find that feedback.'}), 404

@app.route('/changed_password', methods=['GET', 'POST'])
@login_required
def change_password():
    form = updatepasswordForm()  # Form instance for password change

    # Ensure the user is logged in
    if 'username' not in session:
        flash('You must be logged in to change your password.', 'danger')
        return redirect(url_for('login'))

    username = session['username']  # Retrieve the logged-in user's username

    if form.validate_on_submit():
        new_password = form.password.data
        confirm_password = form.confirm_password.data

        if new_password != confirm_password:
            flash('Passwords do not match.', 'danger')
        else:
            try:
                with shelve.open('users.db', 'w') as db:
                    # Check if the user exists in the database
                    if username in db:
                        user_data = db[username]
                        hashed_password = generate_password_hash(new_password, method='pbkdf2:sha256')
                        user_data['password'] = hashed_password
                        db[username] = user_data  # Save updated password
                        flash('Your password has been updated successfully.', 'success')
                        print(user_data['role'])
                        return redirect(url_for('dashboard'))  # Redirect to a post-update page
                    else:
                        flash('User not found. Please log in.', 'danger')
                        return redirect(url_for('login'))
            except Exception as e:
                flash(f'An error occurred: {str(e)}', 'danger')

    return render_template('changed_password.html', form=form)

@app.route('/scheduler/jobs')
@login_required
@role_required('Admin')
def view_scheduled_jobs():
    jobs = scheduler.get_jobs()
    job_list = [{
        'id': job.id,
        'next_run_time': str(job.next_run_time),
        'args': job.args
    } for job in jobs]
    return jsonify(job_list)

@app.route('/retrieve', methods=['GET'])
@login_required
@role_required('Admin')
def retrieve_users():
    page = request.args.get('page', default=1, type=int)  # Get the current page, default to 1
    per_page = 3  # Number of users per page
    search_query = request.args.get('search', default='', type=str).lower()  # Get search input
    with shelve.open('users.db', 'r') as db:
        user_dict = dict(db)  # Convert shelve object to a dictionary
        total_users = len(user_dict)  # Total number of users

        if search_query:
            filtered_users = {
                username: data
                for username, data in user_dict.items()
                if (
                    search_query in username.lower()
                    or search_query in data.get('email', '').lower()
                    or search_query in data.get('role', '').lower()
                    or search_query in data.get('status', '').lower()
                )
            }
        else:
            filtered_users = user_dict
        total_users = len(filtered_users)
        users = list(filtered_users.items())

    # Pagination logic
    start = (page - 1) * per_page
    end = start + per_page
    paginated_users = users[start:end]  # Slice the users for the current page

    total_pages = (total_users + per_page - 1) // per_page  # Calculate total pages
    prev_page = page - 1 if page > 1 else None
    next_page = page + 1 if page < total_pages else None

    # Pass data to the template
    return render_template(
        'retrieve.html',
        users=paginated_users,  # Pass only paginated users
        page=page,
        total_pages=total_pages,
        prev_page=prev_page,
        next_page=next_page,
        total_users=total_users,
        per_page=per_page,
        search_query=search_query,
    )

@app.route('/update/<user_uuid>', methods=['GET', 'POST'])
@login_required
@role_required('Admin')
def update_user(user_uuid):
    form = updateemailrole()
    user_data = None
    username = None

    with shelve_lock:
        with shelve.open('users.db', writeback=True) as db:
            for uname, data in db.items():
                if data.get('uuid') == user_uuid:
                    user_data = data
                    username = uname
                    break

            if not user_data:
                if request.is_json:
                    return jsonify({'message': 'User not found.'}), 404
                flash("User not found.", "danger")
                return redirect(url_for('retrieve_users'))

            # Handle JSON POST from fetch()
            if request.method == 'POST' and request.is_json:
                json_data = request.get_json()

                new_username = json_data.get('username', '').strip()
                new_email = json_data.get('email', '').strip()
                new_role = json_data.get('role')
                new_status = json_data.get('status')

                if not new_username:
                    return jsonify({'message': 'Username cannot be empty.'}), 400
                if new_username != username and new_username in db:
                    return jsonify({'message': 'Username already exists.'}), 400

                user_data['email'] = new_email
                user_data['role'] = new_role if new_role in {'Admin', 'Guest'} else user_data['role']
                user_data['status'] = new_status if new_status in {'Active', 'Suspended'} else user_data['status']

                if new_username != username:
                    db[new_username] = user_data
                    del db[username]
                    if current_user.username == username:
                        session.clear()
                else:
                    db[username] = user_data

                return jsonify({'message': 'User details updated successfully.'}), 200

            form.email.data = user_data.get('email', '')
            form.role.data = user_data.get('role', '')
            form.status.data = user_data.get('status', '')
            return render_template('update_user.html', form=form, user_data=user_data, username=username)

# Delete (Remove User)
@app.route('/delete/<user_uuid>', methods=['POST'])
@login_required
@role_required('Admin')
def delete_user(user_uuid):
    with shelve_lock:
        with shelve.open('users.db', 'w') as db:
            username_to_delete = None
            for uname, user_data in db.items():
                if user_data.get('uuid') == user_uuid:
                    username_to_delete = uname
                    break

            if not username_to_delete:
                return "User not found.", 404
            
            if username_to_delete == session.get('username'):
                return "You cannot delete your own account.", 403

            # Remove scheduler jobs for the user if any
            if scheduler.get_job(f"first_feeding_alert_{user_uuid}") is not None:
                print(f"Removing first feeding alert job for user {username_to_delete}")
            if scheduler.get_job(f"second_feeding_alert_{user_uuid}") is not None:
                print(f"Removing second feeding alert job for user {username_to_delete}")\

            del db[username_to_delete]

    return redirect(url_for('retrieve_users'))

@app.route('/set_ip', methods=['GET', 'POST'])
@login_required
def set_ip():
    global latest_set_ip_settings
    setting = ipForm(request.form)

    if request.method == 'POST':
        if setting.validate():
            try:
                source_ip = setting.source_ip.data
                destination_ip = setting.destination_ip.data
                camera_ip = setting.camera_ip.data
                amcrest_username = setting.amcrest_username.data
                amcrest_password = setting.amcrest_password.data

                try:
                    with shelve.open('settings.db', 'w') as settings_db, shelve.open('IP.db', 'n') as ip_db:
                        port = settings_db.get('Port')
                        if not port:
                            return jsonify(success=False, message="Port setting not found in DB."), 400

                        # Begin packet operations
                        start_syn_packet_thread(source_ip, destination_ip, port, 50000)
                        send_udp_packet(source_ip, "255.255.255.255", 60000, b"\x00\x00\x00\x00\x00")

                        ip_db["IP"] = {
                            "source": source_ip,
                            "destination": destination_ip,
                            "camera_ip": camera_ip,
                            "amcrest_username": amcrest_username,
                            "amcrest_password": amcrest_password
                        }

                    latest_set_ip_settings = {
                        "source": source_ip,
                        "destination": destination_ip,
                        "camera_ip": camera_ip,
                        "amcrest_username": amcrest_username,
                        "amcrest_password": amcrest_password
                    }

                    return jsonify(success=True, redirect_url=url_for('update_setting', mode="auto"))

                except Exception as db_err:
                    print(f"[DB ERROR] {db_err}")
                    return jsonify(success=False, message="Database operation failed."), 500

            except Exception as ex:
                print(f"[PROCESS ERROR] {ex}")
                return jsonify(success=False, message="Unexpected error occurred."), 500

        else:
            # Collect form validation errors
            errors = {field: error for field, error in setting.errors.items()}
            return jsonify(success=False, message="Validation failed.", errors=errors), 400

    return render_template('setIP.html', form=setting)

def get_valid_local_ip(preferred=None):
    interfaces = psutil.net_if_addrs()

    def is_valid(ip):
        return ip and not ip.startswith("127.")

    def is_apipa(ip):
        return ip.startswith("169.254.")

    # 1. Use preferred IP if it exists
    if preferred:
        for iface_addrs in interfaces.values():
            for addr in iface_addrs:
                if addr.family == socket.AF_INET and addr.address == preferred:
                    print(f"Using preferred IP: {preferred}")
                    return preferred

    # 2. Try Wi-Fi or LAN IPs in local devices excluding APIPA
    for iface_name, iface_addrs in interfaces.items():
        for addr in iface_addrs:
            if addr.family == socket.AF_INET and is_valid(addr.address) and not is_apipa(addr.address):
                if "Wi-Fi" in iface_name or "Local Area Connection" in iface_name:
                    print(f"Using fallback interface {iface_name} with IP: {addr.address}")
                    return addr.address

    # 3. Any other valid non-loopback assigned by OS, non-APIPA IP
    for iface_addrs in interfaces.values():
        for addr in iface_addrs:
            if addr.family == socket.AF_INET and is_valid(addr.address) and not is_apipa(addr.address):
                print(f"Using generic fallback IP: {addr.address}")
                return addr.address

    # 4. As last resort, return an APIPA IP if any
    for iface_name, iface_addrs in interfaces.items():
        for addr in iface_addrs:
            if addr.family == socket.AF_INET and is_apipa(addr.address):
                print(f"Using APIPA fallback interface {iface_name} with IP: {addr.address}")
                return addr.address

    raise RuntimeError("No IP address found on any interface")

for iface_name, iface_addrs in psutil.net_if_addrs().items():
    for addr in iface_addrs:
        if addr.family == socket.AF_INET:
            print(f"{iface_name}: {addr.address}")

def send_udp_packet(source, destination, port, message):
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)

    # Try to bind to the user-provided source IP if available
    try:
        sock.bind((source, 0))
        print(f"Socket successfully bound to user-provided source IP: {source}")
    except OSError as e:
        # If binding fails, try to get a valid local IP address in local device
        print(f"Failed to bind to user source IP {source}. Error: {e}")
        try:
            fallback_ip = get_valid_local_ip()
            sock.bind((fallback_ip, 0))
            print(f"Fallback: Socket bound to device IP: {fallback_ip}")
        except Exception as evenin_feed_2:
            # If fallback also fails, bind to any interface which will handle by the OS
            print(f"Fallback failed. Binding to any interface. Error: {evenin_feed_2}")
            sock.bind(("", 0))

    sock.sendto(message, (destination, port))
    print(f"Broadcast sent: {message.hex()} to {destination}:{port}")
    sock.close()

# Define the syn packet function here
def send_syn_packet_in_background(client_ip, server_ip, source_port, destination_port):
    try:
        # Call the send_syn_packet function in a new thread
        send_syn_packet(client_ip, server_ip, source_port, destination_port)
    except Exception as e:
        print(f"Error in background thread: {e}")

def start_syn_packet_thread(client_ip, server_ip, port, destination_port):
    """Run the send_syn_packet in the background"""
    syn_thread = threading.Thread(target=send_syn_packet_in_background, args=(client_ip, server_ip, port, destination_port))
    syn_thread.daemon = True  # Allow the thread to exit when the main program exits
    syn_thread.start()

def initialize_databases():
    print("Creating new databases and initializing default values.")
    # settings.db
    with shelve.open('settings.db', 'c') as db:
        db['Generate_Status'] = False
        email_setup = Email('iatfadteam@gmail.com', 'iatfadteam@gmail.com', 'pmtu cilz uewx xqqi', 3)
        db['Email_Data'] = {'Email_Info': email_setup}
        db['Port'] = 53101
        db['last_ip_id'] = 7500
        db['Confidence_Rate'] = 60 
        db['syn_ack_seq'] = None
        db['syn_ack_ack'] = None    

def update_existing_databases():
    print("Updating existing database values if needed.")

    with shelve.open('users.db', 'c') as db:
        for key in db:
            user_data = db[key]
            if 'status' not in user_data:
                user_data['status'] = 'Active'
                db[key] = user_data

    with shelve.open('settings.db', 'w') as db:
        db['Port'] = 53101
        db['last_ip_id'] = 7500

def setup_mail():
    print("Configuring mail...")
    app.config['MAIL_USERNAME'] = 'iatfadteam@gmail.com'
    app.config['MAIL_PASSWORD'] = 'pmtu cilz uewx xqqi'
    app.config['MAIL_DEFAULT_SENDER'] = ('Admin', 'iatfadteam@gmail.com')
    return Mail(app)

camera_threads = []
camera_lock = threading.Lock()

def start_camera_threads():
    global camera_threads
    with camera_lock:
        if camera_threads:
            return camera_threads

        camera_stop_event.clear()

        capture_thread = threading.Thread(target=capture_frames, daemon=True)
        video_thread = threading.Thread(target=video_processing_loop, daemon=True)

        capture_thread.start()
        time.sleep(5)
        video_thread.start()

        camera_threads = [capture_thread, video_thread]
        print("Camera threads started.")
        return camera_threads

def stop_camera_threads():
    global camera_threads, freshest_frame
    with camera_lock:
        camera_stop_event.set()

        with freshest_frame_lock:
            freshest_frame.stop(5)
            freshest_frame = None

        for thread in camera_threads:
            if thread.is_alive():
                thread.join(timeout=5)

        camera_threads = []
        camera_stop_event.clear()
        print("Camera thread stopped.")

def start_threads():
    print("Starting all system threads...")

    feeding_thread = threading.Thread(target=feeding_scheduler_loop) # Schedule feeding tasks based on the configured times
    validate_thread = threading.Thread(target=validate_config_thread) # Validate the configuration and settings periodically

    feeding_thread.start()
    validate_thread.start()
    reschedule_feeding_alerts()

    return [feeding_thread, validate_thread]


def cleanup_on_exit():
    print("Cleaning up and closing ports...")
    stop_event.set()

    try:
        with shelve.open('settings.db', 'r') as db:
            port = db.get('Port')
            server_isn = db.get('syn_ack_seq')
            server_ack = db.get('syn_ack_ack')

        with shelve.open("IP.db") as db:
            ip_data = db.get("IP", {})
            source_ip = ip_data.get("source", "N/A")
            destination_ip = ip_data.get("destination", "N/A")

        send_RST_packet(port, server_isn, server_ack, source_ip, destination_ip)
        print("Port closed successfully.")
    except Exception as e:
        print(f"[CLEANUP ERROR] Failed to send RST packet or close ports: {e}")

if __name__ == '__main__':
    try:
        # Check if DB is readable
        with shelve.open('settings.db', 'r') as _:
            pass
        update_existing_databases()

    except Exception as e:
        print(f"[INIT ERROR] Could not open DB, initializing new one: {e}")
        initialize_databases()

    # Setup mail
    mail = setup_mail()
    
    #Feedback Store
    store_Feedback = FeedbackStore()

    # Start all system threads
    threads = start_threads()

    # Start Flask app
    try:
        app.run(host='0.0.0.0', port=5001, debug=False)
    finally:
        cleanup_on_exit()
        for t in threads:
            t.join()